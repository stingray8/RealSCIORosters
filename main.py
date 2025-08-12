import pulp
from pulp import LpProblem, LpVariable, LpMaximize, PULP_CBC_CMD
import time
import copy
import math
import pandas as pd
import openpyxl
import numpy as np
from tabulate import tabulate
from fuzzywuzzy import fuzz
import sys
from openpyxl import load_workbook
import Levenshtein
from scipy.stats import norm
from pulp import value
from functions import *
import gc
gc.enable()

# File path
file_path = 'Compute Rosters.xlsx'

team_df = pd.read_excel(file_path, sheet_name="Team")
team_info = data_frame_to_np(team_df)
TEAM_MEMBERS = team_info[:, 0].tolist()
TEAM_MEMBERS.pop(0)
TEAM_MEMBERS = [item.title() for item in TEAM_MEMBERS if isinstance(item, str) or not math.isnan(item)]
del team_df

# If TEAM_MEMBERS is full, only those people will be used.
# If TEAM_MEMBERS is partially full, the algorithm will pick the next best people.
RESTRICTED_FROM_TEAM = team_info[:, 1].tolist()
RESTRICTED_FROM_TEAM.pop(0)
RESTRICTED_FROM_TEAM = [item.title() for item in RESTRICTED_FROM_TEAM if isinstance(item, str) or not math.isnan(item)]
RESTRICTED_FROM_TEAM = set(RESTRICTED_FROM_TEAM)

for person in TEAM_MEMBERS:
    if person in RESTRICTED_FROM_TEAM:
        raise Exception(f"{person} is in Team List and is also restricted")

NO_SAME_PAIRS = team_info[:, 4].tolist()
NO_SAME_PAIRS = NO_SAME_PAIRS[1]
if 'T' in NO_SAME_PAIRS:
    NO_SAME_PAIRS = False
    print("Same partners on different events is allowed")
else:
    NO_SAME_PAIRS = True
    print("Same partners on different events is not allowed")
    print("This is ignored for three person events")

check_spelling = team_info[:, 6].tolist()
if 'T' in check_spelling:
    check_spelling = True
    print("Checking spelling")
else:
    check_spelling = False
    print_red("Will not check spelling")

MUST_INCLUDE = {}
if 0 <= len(TEAM_MEMBERS) < 15:
    MUST_INCLUDE = set(TEAM_MEMBERS)
    print_red("Not enough people. Filling " + str(15 - len(TEAM_MEMBERS)) + " spot(s)")
    TEAM_MEMBERS = []

elif len(TEAM_MEMBERS) > 15:
    print_red("Too many people, picking best 15")

schedule = team_info[1][7]
print(f"Schedule: {schedule}")

print("------------")
name_age = team_info[:, 2:4]
name_age = name_age[1:]
name_age = [
    [str(item[0]).title(), item[1]]
    for item in name_age
    if isinstance(item[0], str) or not math.isnan(item[0])
]

# Read event ratings
event_ratings_df = pd.read_excel(file_path, sheet_name="Event Ratings")
event_ratings = data_frame_to_np(event_ratings_df).tolist()
del event_ratings_df
events = event_ratings[0]
# Remove the first element (assumed to be a label like "Name")
events.pop(0)
event_ratings.pop(0)
events = [e.title() for e in events]
events = tuple(events)
num_events = len(events)

# Read past performance data
past_performance_df = pd.read_excel(file_path, sheet_name="Past Performance")
past_performance = data_frame_to_np(past_performance_df).tolist()
del past_performance_df
past_performance.pop(0)
past_performance = [
    [str(item[0]).title(), str(item[1]).title(), *item[2:]]  # Title case first two columns, keep rest unchanged
    for item in past_performance
    if (isinstance(item[0], str) or not math.isnan(item[0]))  # Filter valid entries
]

finetune = pd.read_excel(file_path, sheet_name="Finetune")
finetune = data_frame_to_np(finetune)

tournament_weights = finetune[:, 0].tolist()
tournament_weights.pop(0)

# Insert weights based on performance level
for i in range(len(past_performance)):
    if not math.isnan(past_performance[i][3]):  # Middle Regionals
        past_performance[i].insert(3, tournament_weights[0])
    elif not math.isnan(past_performance[i][4]):  # Middle States
        past_performance[i].insert(3, tournament_weights[1])
    elif not math.isnan(past_performance[i][5]):  # Middle Nationals
        past_performance[i].insert(3, tournament_weights[2])
    elif not math.isnan(past_performance[i][6]):  # High Regionals
        past_performance[i].insert(3, tournament_weights[3])
    elif not math.isnan(past_performance[i][7]):  # High States
        past_performance[i].insert(3, tournament_weights[4])
    elif not math.isnan(past_performance[i][8]):  # High Nationals
        past_performance[i].insert(3, tournament_weights[5])
del tournament_weights
past_performance = [row[:4] + [row[10]] for row in past_performance]
# Read time conflicts
try:
    time_conflicts_df = pd.read_excel(file_path, sheet_name=schedule)
except ValueError:
    raise Exception("Schedule sheet not found.")

print_logs = 'T' in team_info[1][8]
print(f"printing logs == {print_logs}")

time_conflicts_df = time_conflicts_df.replace('nan', np.nan)
time_conflicts = data_frame_to_np(time_conflicts_df)
time_conflicts = [[item for item in row if not pd.isna(item)] for row in time_conflicts]
time_conflicts = [[item.title() for item in row if "Unnamed" not in str(item)] for row in time_conflicts]
time_conflicts = [tuple(row) for row in time_conflicts]

# Read events that require three people
events_with_three_people_df = pd.read_excel(file_path, sheet_name="3 Person Events")
events_with_three_people = data_frame_to_np(events_with_three_people_df).tolist()
events_with_three_people = tuple(event[0].title() for event in events_with_three_people)

# Read extra info
extra_info_df = pd.read_excel(file_path, sheet_name="Extra Info")
extra_info = data_frame_to_np(extra_info_df)

# Process "Do Not Work Well Together" (first two columns)
do_not_work_well_together = extra_info[:, 0:2].tolist()
do_not_work_well_together.pop(0)
do_not_work_well_together = [
    [item[0].title(), item[1].title()]
    for item in do_not_work_well_together
    if (isinstance(item[0], str) or not math.isnan(item[0])) and
       (isinstance(item[1], str) or not math.isnan(item[1]))
]

must_have_events = extra_info[:, 2:4].tolist()
must_have_events.pop(0)
must_have_events = [
    tuple([str(item[0]).title(), str(item[1]).title()])
    for item in must_have_events
    if isinstance(item[0], str) or not math.isnan(item[0])
]

# Process "People With Three" (column 4)
people_with_three = extra_info[:, 4].tolist()
people_with_three.pop(0)
people_with_three = [item.title() for item in people_with_three if isinstance(item, str) or not math.isnan(item)]

# Process "People Cannot Participate in Event" (columns 5 and 6)
people_cannot_event = extra_info[:, 5:7].tolist()
people_cannot_event.pop(0)
people_cannot_event = [
    tuple([str(item[0]).title(), str(item[1]).title()])
    for item in people_cannot_event
    if isinstance(item[0], str) or not math.isnan(item[0])
]


# Process "Should Work Together"
work_together = extra_info[:, 11:13].tolist()
work_together.pop(0)  # Remove header
work_together = [
    tuple([item[0].title(), item[1].title()])
    for item in work_together
    if isinstance(item[0], str) and isinstance(item[1], str)
]

people_cannot_category = extra_info[:, 7:9].tolist()
people_cannot_category.pop(0)  # Remove header
people_cannot_category = [
    tuple([item[0].title(), item[1].title()])
    for item in people_cannot_category
    if isinstance(item[0], str) and isinstance(item[1], str)
]

event_to_category = dict()
category_to_events = dict()
category_info = pd.read_excel(file_path, sheet_name="Event Categories")
category_info = data_frame_to_np(category_info)
category_info = category_info.T
for c in range(len(category_info)):
    for a in range(len(category_info[c])):
        category_info[c][a] = str(category_info[c][a]).title()

for row in category_info:
    row = row.tolist()
    section = row[0]
    category_to_events[section] = []
    for i in range(1, len(row)):
        if not row[i] == "Nan":
            event_to_category[row[i]] = section
            category_to_events[section].append(row[i])

event_exceptions = extra_info[:, 9:11].tolist()
event_exceptions.pop(0)  # Remove header
event_exceptions = [
    tuple([item[0].title(), item[1].title()])
    for item in event_exceptions
    if isinstance(item[0], str) and isinstance(item[1], str)
]


if 'T' in finetune[:, 6][1]:
    print("Normalizing event ratings to have same mean=5")
else:
    print_red("Not normalizing event ratings to have same mean")


class Person:
    def __init__(self, data, events):
        self.name = data[0].strip().title()

        data.pop(0)

        nan_values = False
        for r in range(len(data)):
            if str(data[r]) == "nan":
                nan_values = True
                data[r] = 0
        if nan_values:
            print_red("NaN values were found for " + str(self.name) + ". Replacing with zeroes")
        data = list(map(int, data))

        self.events = events
        self.original_ratings = {events[i]: data[i] for i in range(len(data))}
        if 'T' in finetune[:, 6][1]:
            data = normalize_ratings(data)

        self.ratings = {events[i]: data[i] for i in range(len(data))}
        self.age = None

    def get_all_scores(self):
        return [self.ratings[event] for event in self.events]

    def replace_event_rating(self, event, score):
        self.ratings[event] = score

    def get_event_rating(self, event):
        return self.ratings[event]

    def get_original_rating(self, event):
        return self.original_ratings[event]

    def __str__(self):
        ratings_str = ", ".join([f"{event}: {self.ratings[event]}" for event in self.events])
        return f"Name: {self.name, self.age}\nEvents and Ratings:\n{ratings_str}"


# Build the People List
people = []
people_dict = {}
all_people_set = set()

names = [row[0] for row in event_ratings]

print("------")

for p in MUST_INCLUDE:
    if p not in names:
        raise Exception(f"Must include {p} who doesn't exist")

for row in event_ratings:
    name = row[0].strip().title()
    all_people_set.add(name)

    # In provided-team mode, only include names from TEAM_MEMBERS.
    if ((not TEAM_MEMBERS) or (name in TEAM_MEMBERS)) and (name not in RESTRICTED_FROM_TEAM) or (
            len(TEAM_MEMBERS) > 15 and name in TEAM_MEMBERS):
        person = Person(row, events)
        people.append(person)
        people_dict[name] = person

num_people = len(people)


if check_spelling:
    spelling_threshold = int(finetune[:, 3].tolist()[1])
    print("Checking for name mistakes...")
    # Check for possible name misspellings
    for name in all_people_set:
        if TEAM_MEMBERS:
            for team_name in TEAM_MEMBERS:
                if team_name != name and calculate_string_similarity(team_name, name) > spelling_threshold:
                    print_red(f"Possible spelling inconsistency in Team Names: '{team_name}' and form name '{name}'")

        for performance in past_performance:
            performance_name = performance[0]
            if performance_name != name and calculate_string_similarity(performance_name, name) > spelling_threshold:
                print_red(
                    f'Possible spelling inconsistency in past performance: "{performance_name}" and form name "{name}"')

        for pair in RESTRICTED_FROM_TEAM:
            if pair != name and calculate_string_similarity(pair, name) > spelling_threshold:
                print_red(f"Possible spelling inconsistency in restricted from team: '{pair}' and form name '{name}'")

        for pair in do_not_work_well_together:
            if pair[0] != name and calculate_string_similarity(pair[0], name) > spelling_threshold:
                print_red(f"Possible spelling inconsistency in poor synergy: '{pair[0]}' and form name '{name}'")
            if pair[1] != name and calculate_string_similarity(pair[1], name) > spelling_threshold:
                print_red(f"Possible spelling inconsistency in poor synergy: '{pair[1]}' and form name '{name}'")

        for pair in must_have_events:
            if pair[0] != name and calculate_string_similarity(pair[0], name) > spelling_threshold:
                print_red(f"Possible spelling inconsistency in must have events: '{pair[0]}' and form name '{name}'")

        for p in people_with_three:
            if p != name and calculate_string_similarity(p, name) > spelling_threshold:
                print_red(f"Possible spelling inconsistency in people with three events: '{p}' and form name '{name}'")

        for pair in people_cannot_event:
            if pair[0] != name and calculate_string_similarity(pair[0], name) > spelling_threshold:
                print_red(
                    f"Possible spelling inconsistency in people restricted from events: '{pair[0]}' and form name '{name}'")

        for pair in work_together:
            if pair[0] != name and calculate_string_similarity(pair[0], name) > spelling_threshold:
                print_red(
                    f"Possible spelling inconsistency in should work together: '{pair[0]}' and form name '{name}'")
            if pair[1] != name and calculate_string_similarity(pair[1], name) > spelling_threshold:
                print_red(
                    f"Possible spelling inconsistency in should work together: '{pair[1]}' and form name '{name}'")

        for pair in name_age:
            if pair[0] != name and calculate_string_similarity(pair[0], name) > spelling_threshold:
                print_red(f"Possible spelling inconsistency in name age columns: '{pair[0]}' and form name '{name}'")
        for pair in people_cannot_category:
            if pair[0] != name and calculate_string_similarity(pair[0], name) > spelling_threshold:
                print_red(
                    f"Possible spelling inconsistency in people restricted from events: '{pair[0]}' and form name '{name}'")
        for pair in event_exceptions:
            if pair[0] != name and calculate_string_similarity(pair[0], name) > spelling_threshold:
                print_red(
                    f"Possible spelling inconsistency in people restricted from events: '{pair[0]}' and form name '{name}'")

    print("Checking for event name mistakes...")
    for event in events:
        for performance in past_performance:
            performance_score = performance[1]
            if performance_score != event and calculate_string_similarity(performance_score,
                                                                          event) > spelling_threshold:
                print_red(
                    f'Possible spelling inconsistency in past performance: "{performance_score}" and form name "{event}"')

        for row in time_conflicts:
            for item in row:
                if item != event and calculate_string_similarity(item, event) > spelling_threshold:
                    print_red(f"Possible spelling inconsistency in time conflicts: '{item}' and form name '{event}'")

        for pair in must_have_events:
            if pair[1] != event and calculate_string_similarity(pair[1], event) > spelling_threshold:
                print_red(f"Possible spelling inconsistency in must have events: '{pair[1]}' and form name '{event}'")

        for pair in people_cannot_event:
            if pair[1] != event and calculate_string_similarity(pair[1], event) > spelling_threshold:
                print_red(
                    f"Possible spelling inconsistency in people restricted from events: '{pair[1]}' and form name '{event}'")

        for section in category_info:
            for e in section:
                e = str(e)
                if e != event and calculate_string_similarity(e, event) > spelling_threshold:
                    print_red(f"Possible spelling inconsistency in event categories: '{e}' and form name '{event}'")
        for pair in people_cannot_category:
            if pair[1] != event and calculate_string_similarity(pair[1], event) > spelling_threshold:
                print_red(
                    f"Possible spelling inconsistency in people restricted from category: '{pair[0]}' and form name '{name}'")
        for pair in event_exceptions:
            if pair[1] != event and calculate_string_similarity(pair[1], event) > spelling_threshold:
                print_red(
                    f"Possible spelling inconsistency in people-event exceptions: '{pair[0]}' and form name '{name}'")

    for category in category_info:
        for pair in people_cannot_category:
            if pair[1] != category[0] and calculate_string_similarity(pair[1], category[0]) > spelling_threshold:
                print_red(
                    f"Possible spelling inconsistency in people restricted from categories: '{pair[1]}' and category name '{category[0]}'")


event_exceptions = set(event_exceptions)
for p,e in people_cannot_event:
    if (p,e) in event_exceptions:
        raise Exception(f"{p} in event {e} in both event restricted (not category) and event exception")

set_people_cannot_event = set(people_cannot_event)
for row in people_cannot_category:
    p = row[0]
    e = row[1]
    try:
        for i in range(len(category_to_events[e])):
                if ((p, category_to_events[e][i]) not in event_exceptions) and ((p, category_to_events[e][i]) not in set_people_cannot_event):
                    people_cannot_event.append((p, category_to_events[e][i]))
    except KeyError as err:
        print_red(f"There are spelling errors for {e} in people restricted from category for person {p}")
        print_red("\tSkipping restriction for now")

del event_exceptions
del set_people_cannot_event
del category_to_events

gc.collect()

print("Cleaning lists...")
# Clean Up Constraint Lists
must_have_events = [item for item in must_have_events if item[0] in people_dict]
people_with_three = [item for item in people_with_three if item in people_dict]
do_not_work_well_together = [pair for pair in do_not_work_well_together if
                             pair[0] in people_dict and pair[1] in people_dict]
work_together = [pair for pair in work_together if pair[0] in people_dict and pair[1] in people_dict]

people_cannot_event = [item for item in people_cannot_event if item[0] in people_dict]


# Set up name_age
name_age = [item for item in name_age if item[0] in people_dict]
for pair in name_age:
    people_dict[pair[0]].age = pair[1]

# Clean past_performance
i = 0
while i < len(past_performance):

    # Giving a better baseline scores based on number of past events
    if past_performance[i][0] in people_dict:
        person_name = people_dict[past_performance[i][0]]
        for event_key in person_name.ratings.keys():
            try:
                if event_to_category[event_key] == event_to_category[past_performance[i][1]] and not event_key == \
                                            past_performance[i][1] and not event_to_category[event_key] == "Extra":
                    if print_logs:
                        print(past_performance[i])
                        print(f"Same category as {event_key}. Current rating is {person_name.ratings[event_key]}")
                    person_name.ratings[event_key] += max(0, find_placement_score(past_performance[i][2]) * \
                                                          finetune[:, 5].tolist()[1] * past_performance[i][3])
                    person_name.ratings[event_key] = round(person_name.ratings[event_key], 2)
                    if print_logs:
                        print(f"New rating is {person_name.ratings[event_key]} \n")
            except KeyError as e:
                raise Exception(f"KeyError: Missing event '{e.args[0]}' in event categories")

    if past_performance[i][0] not in people_dict or past_performance[i][1] not in events or past_performance[i][
        0] in RESTRICTED_FROM_TEAM:
        if past_performance[i][0] not in people_dict:
            reason = "Not in people list"
        else:
            reason = "Event not in event list"
        if print_logs:
            print(f"Removed from past_performance: {past_performance[i]} {reason}")

        past_performance.pop(i)
        i -= 1
    i += 1

if TEAM_MEMBERS:
    do_not_work_well_together = [
        pair for pair in do_not_work_well_together
        if pair[0] in TEAM_MEMBERS and pair[1] in TEAM_MEMBERS
    ]

    work_together = [
        pair for pair in work_together
        if pair[0] in TEAM_MEMBERS and pair[1] in TEAM_MEMBERS
    ]

do_not_work_well_together = [
    pair for pair in do_not_work_well_together
    if pair[0] not in RESTRICTED_FROM_TEAM and pair[1] not in RESTRICTED_FROM_TEAM
]

work_together = [
    pair for pair in work_together
    if pair[0] not in RESTRICTED_FROM_TEAM and pair[1] not in RESTRICTED_FROM_TEAM
]
work_together_bonus_weight = finetune[:, 1][9]
if work_together_bonus_weight != 0:
    seen = set()

    filtered = [
        pair for pair in work_together
        if (key := tuple(sorted(pair))) not in seen and not seen.add(key)
    ]

    work_together = filtered
else:
    work_together = []



highest_score_counted = finetune[:, 2].tolist()[1]
current_year = team_info[:, 9].tolist()[1]

checked = []
for past in reversed(past_performance):
    person_name, event_name, skill_level, weight, year = past
    if skill_level > highest_score_counted:
        continue
    if skill_level > highest_score_counted:
        pass
    if person_name not in all_people_set:
        raise Exception(f"{person_name} not found")
    if person_name in people_dict:
        person = people_dict[person_name]
        current_rating = person.get_event_rating(event_name)
        if year >= current_year - 1 and (person_name, event_name) not in checked:
            new_rating = current_rating + (weight * find_placement_score(skill_level))

        else:
            new_rating = current_rating + (finetune[:, 1].tolist()[1] * (weight * find_placement_score(skill_level)))
        if find_placement_score(1) * weight + 10 < new_rating:
            new_rating = find_placement_score(1) * weight + 10
        person.replace_event_rating(event_name, new_rating)
        checked.append((person_name, event_name))
del past_performance
if print_logs:
    for p in people:
        print(p)
# Create Mappings for the Optimization Model
event_to_index = {event: idx for idx, event in enumerate(events)}
index_to_event = {idx: event for event, idx in event_to_index.items()}
person_to_index = {person.name: idx for idx, person in enumerate(people)}
index_to_person = {idx: person.name for idx, person in enumerate(people)}

start_time = time.time()
output_target = open("output.txt", "a")


def get_people():
    return people


# Create LP problem
if 'T' in finetune[:, 2].tolist()[9]:
    print_red("Creating worst possible team.")
    prob = pulp.LpProblem("Maximize_Event_Assignment_Score", pulp.LpMinimize)
    time.sleep(1)
else:
    prob = pulp.LpProblem("Maximize_Event_Assignment_Score", pulp.LpMaximize)

print("Defining problem")
if len(TEAM_MEMBERS) > 15:
    TEAM_MEMBERS = []

if TEAM_MEMBERS:
    # Provided team mode: only assignment variables.
    x = pulp.LpVariable.dicts("x", (range(num_people), range(num_events)), cat='Binary')
else:
    # Choose-best mode: assignment variables and selection variables.
    x = pulp.LpVariable.dicts("x", (range(num_people), range(num_events)), cat='Binary')
    y = pulp.LpVariable.dicts("y", range(num_people), cat='Binary')

    for i, person in enumerate(people):
        if person.name in MUST_INCLUDE:
            prob += y[i] == 1, f"Must_include_{person.name}"

# Objective Function
# Base score from event ratings


objective = pulp.lpSum(people[i].get_all_scores()[j] * x[i][j]
                       for i in range(num_people) for j in range(num_events))


min_preference_score = finetune[:, 3].tolist()[9]
if str(min_preference_score) == 'nan':
    min_preference_score=0
    print_red("No minimum preference score set. Defaulting to 0")

for i in range(num_people):
    for j in range(num_events):
        if people[i].get_original_rating(index_to_event[j]) < min_preference_score:
            prob += x[i][j] == 0, f"No_Low_Score2_{i}_{j}"

MIN_EVENT_SCORE = finetune[:, 4].tolist()[9]
if str(MIN_EVENT_SCORE) == 'nan':
    MIN_EVENT_SCORE = 0
    print_red("No minimum preference score set. Defaulting to 0")

for j in range(num_events):
    prob += pulp.lpSum(people[i].get_event_rating(index_to_event[j]) * x[i][j]
                       for i in range(num_people)) >= MIN_EVENT_SCORE, f"MinScore_{j}"


z = pulp.LpVariable.dicts("work_together_bonus",
                          [(person_to_index[p1], person_to_index[p2], j)
                           for (p1, p2) in work_together
                           for j in range(num_events)
                           if p1 in person_to_index and p2 in person_to_index],
                          cat='Binary')

for (p1, p2) in work_together:
    if p1 in person_to_index and p2 in person_to_index:
        p1_idx = person_to_index[p1]
        p2_idx = person_to_index[p2]

        for j in range(num_events):
            # z <= x1
            prob += z[(p1_idx, p2_idx, j)] <= x[p1_idx][j], f"work_bonus_z_leq_x1_{p1}_{p2}_{j}"
            # z <= x2
            prob += z[(p1_idx, p2_idx, j)] <= x[p2_idx][j], f"work_bonus_z_leq_x2_{p1}_{p2}_{j}"
            # z >= x1 + x2 - 1
            prob += z[(p1_idx, p2_idx, j)] >= x[p1_idx][j] + x[p2_idx][j] - 1, f"work_bonus_z_geq_{p1}_{p2}_{j}"

# Add the bonus term to having pairs that want to work with each other
objective += work_together_bonus_weight * pulp.lpSum(
    z[(person_to_index[p1], person_to_index[p2], j)]
    for (p1, p2) in work_together
    for j in range(num_events)
    if p1 in person_to_index and p2 in person_to_index
)

prob += objective, "Total_Score"

# Linking & Team Size Constraints (Choose-Best Mode)
if not TEAM_MEMBERS:
    prob += pulp.lpSum(y[i] for i in range(num_people)) == 15, "Team_Size"
    for i in range(num_people):
        for j in range(num_events):
            prob += x[i][j] <= y[i], f"Link_Assign_Select_{i}_{j}"

# Time Conflict Constraints
for conflict_set in time_conflicts:
    conflict_indices = [event_to_index[event] for event in conflict_set if event in event_to_index]
    for i in range(num_people):
        prob += pulp.lpSum(x[i][j] for j in conflict_indices) <= 1, f"Time_Conflict_Person_{i}_{conflict_set}"

max_age_12 = team_info[:, 5].tolist()[1]
age_12_indices = [i for i, person in enumerate(people) if person.age == 12]

if TEAM_MEMBERS:
    # If in team mode, everyone is already in the team, so use x variables to check involvement
    prob += pulp.lpSum(
        x[i][j] for i in age_12_indices for j in range(num_events)) <= max_age_12 * num_events, "Max_Age_12_Team"
else:
    # If choosing team, use y variables to control who is selected
    prob += pulp.lpSum(y[i] for i in age_12_indices) <= max_age_12, "Max_Age_12_Select"

# No Same Pairs Constraint (Optional)
if NO_SAME_PAIRS:
    for i in range(num_people):
        for j in range(i + 1, num_people):
            for event1 in range(num_events):
                if index_to_event[event1] in events_with_three_people:
                    continue
                for event2 in range(event1 + 1, num_events):
                    if index_to_event[event2] in events_with_three_people:
                        continue
                    prob += x[i][event1] + x[j][event1] + x[i][event2] + x[j][event2] <= 3, \
                        f"No_Same_Pairs_{i}_{j}_{event1}_{event2}"

# People Cannot Participate in Certain Events
for person_name, event_name in people_cannot_event:
    if person_name in person_to_index and event_name in event_to_index:
        p_idx = person_to_index[person_name]
        e_idx = event_to_index[event_name]
        prob += x[p_idx][e_idx] == 0, f"Cannot_Participate_{person_name}_{event_name}"

# Must-Have Events Constraint
for person_name, event_name in must_have_events:
    if person_name in person_to_index and event_name in event_to_index:
        p_idx = person_to_index[person_name]
        e_idx = event_to_index[event_name]
        if TEAM_MEMBERS:
            prob += x[p_idx][e_idx] == 1, f"Must_Have_{person_name}_{event_name}"
        else:
            prob += x[p_idx][e_idx] == y[p_idx], f"Must_Have_{person_name}_{event_name}"

# --- People who don't work well together: never assigned together on the same event ---
if do_not_work_well_together:
    for pair in do_not_work_well_together:
        person1_idx = person_to_index[pair[0]]
        person2_idx = person_to_index[pair[1]]
        for event_idx in range(num_events):
            prob += x[person1_idx][event_idx] + x[person2_idx][event_idx] <= 1, f"Do_Not_Work_Well_{pair}_{event_idx}"

# Difference in Number of Events Constraint
if 'T' in finetune[:, 4].tolist()[1]:
    print("Balancing team distribution")
    if TEAM_MEMBERS:
        for i in range(num_people - 1):
            for j in range(i + 1, num_people):
                prob += pulp.lpSum(x[i][k] for k in range(num_events)) - \
                        pulp.lpSum(x[j][k] for k in range(num_events)) <= 1, f"Diff_{i}_{j}_upper"
                prob += pulp.lpSum(x[j][k] for k in range(num_events)) - \
                        pulp.lpSum(x[i][k] for k in range(num_events)) <= 1, f"Diff_{j}_{i}_upper"
    else:
        M = num_events
        for i in range(num_people - 1):
            for j in range(i + 1, num_people):
                prob += (pulp.lpSum(x[i][k] for k in range(num_events)) -
                         pulp.lpSum(x[j][k] for k in range(num_events))
                         ) <= 1 + (2 - y[i] - y[j]) * M, f"Diff_{i}_{j}_upper"
                prob += (pulp.lpSum(x[j][k] for k in range(num_events)) -
                         pulp.lpSum(x[i][k] for k in range(num_events))
                         ) <= 1 + (2 - y[i] - y[j]) * M, f"Diff_{j}_{i}_upper"
else:
    print_red("Not balancing team distribution")

# Event Assignment Constraint
# Each event must get the required number of people.
for j in range(num_events):
    if index_to_event[j] in events_with_three_people:
        prob += pulp.lpSum(x[i][j] for i in range(num_people)) == 3, f"Event_{j}_Exactly_3"
    else:
        prob += pulp.lpSum(x[i][j] for i in range(num_people)) == 2, f"Event_{j}_Exactly_2"

# People With Three Constraint
for person_name in people_with_three:
    if person_name in person_to_index:
        p_idx = person_to_index[person_name]
        prob += pulp.lpSum(x[p_idx][j] for j in range(num_events)) <= 3, f"Max_Three_{person_name}"


def solve():
    # Solve the Model
    print("Solving")
    prob.solve(PULP_CBC_CMD(msg=print_logs))
    model_score = value(prob.objective)
    with open("output.txt", "w"):
        pass

    status = pulp.LpStatus[prob.status]
    print("Status:", status, file=output_target)
    if status == "Infeasible":
        print_red("No feasible solution found")

    else:
        # Extract and Output Assignments
        assignments = []
        if TEAM_MEMBERS:
            for i in range(num_people):
                assigned_events = [index_to_event[j] for j in range(num_events) if pulp.value(x[i][j]) == 1]
                assignments.append((people[i].name, assigned_events))
        else:
            for i in range(num_people):
                if pulp.value(y[i]) == 1:
                    assigned_events = [index_to_event[j] for j in range(num_events) if pulp.value(x[i][j]) == 1]
                    assignments.append((people[i].name, assigned_events))
        get_results(assignments, model_score=model_score)
        return assignments


def get_results(assignments, model_score=None, event_assignments=None):
    # Print assignments
    print("Assignments:", file=output_target)
    assignment_array = []
    for person_name, event_list in assignments:
        line = [person_name] + event_list
        assignment_array.append(line)
    assignment_array.sort(key=lambda x: x[0])
    print(tabulate(assignment_array), file=output_target)
    print(file=output_target)

    # Build event assignments dictionary
    if event_assignments is None:
        event_assignments = {event: [] for event in events}
        if TEAM_MEMBERS:
            for i in range(num_people):
                for j in range(num_events):
                    if pulp.value(x[i][j]) == 1:
                        event_assignments[index_to_event[j]].append(people[i].name)
        else:
            for i in range(num_people):
                if pulp.value(y[i]) == 1:
                    for j in range(num_events):
                        if pulp.value(x[i][j]) == 1:
                            event_assignments[index_to_event[j]].append(people[i].name)
    print("Event Assignments:", file=output_target)
    event_array = []
    for event, assigned_people in event_assignments.items():
        line = [event] + assigned_people
        event_array.append(line)
    event_array.sort(key=lambda x: x[0])
    print(tabulate(event_array, headers=["Event", "Person 1", "Person 2", "Person 3"]), file=output_target)
    print(file=output_target)

    print("Event Assignments with scores:", file=output_target)
    event_array = []
    for event, assigned_people in event_assignments.items():
        line = [event] + assigned_people

        org_sco = 0
        cur_sco = 0
        for person in assigned_people:
            org_sco += people_dict[person].get_original_rating(event)
            cur_sco += people_dict[person].get_event_rating(event)
        if len(line) < 4:
            line.append("")
        org_sco /= len(assigned_people)
        cur_sco /= len(assigned_people)
        line.append(org_sco)
        line.append(cur_sco)
        event_array.append(line)
    event_array.sort(key=lambda x: x[0])
    print(tabulate(event_array,
                   headers=["Event", "Person 1", "Person 2", "Person 3", "Avg. orig. score", "Avg. adj. score"]),
          file=output_target)
    print(file=output_target)

    # Calculate total score and individual statistics
    total_score = 0
    original_score = 0
    individual_stats = []

    for person_name, event_list in assignments:
        if not event_list:
            continue
        person_score = 0
        person_happiness = 0
        event_details_list = []  # Will hold strings like "Event (Orig: X, Curr: Y)"
        for event in event_list:
            current_rating = people_dict[person_name].get_event_rating(event)
            original_rating = people_dict[person_name].get_original_rating(event)
            person_score += current_rating
            person_happiness += original_rating
            total_score += current_rating
            original_score += original_rating
            event_details_list.append((event, original_rating, current_rating))
        avg_score = round(person_score / len(event_list), 2)
        avg_happiness = round(person_happiness / len(event_list), 2)
        event_details_list.sort(key=lambda x: x[1], reverse=True)
        event_details_list = [f"{tup[0]} ({tup[1]}, {tup[2]})" for tup in event_details_list]
        add_to_stats = [person_name, avg_happiness, avg_score] + event_details_list
        individual_stats.append(add_to_stats)
    individual_stats.sort(key=lambda x: x[0])
    from itertools import combinations

    work_together_set = {tuple(sorted(pair)) for pair in work_together}

    for event, assigned_people in event_assignments.items():
        for pair in combinations(sorted(assigned_people), 2):
            if pair in work_together_set:
                total_score += work_together_bonus_weight

    print("Individual Stats with Event Details:", file=output_target)
    print(tabulate(individual_stats, headers=["Name", "Happiness", "Score", "Event1", "Event2", "Event3", "Event 4"], ),
          file=output_target)

    print("Team Created")
    print("------------")
    print(f"Preference Score: {original_score}")
    print(
        f"Happiness Percent: {int(original_score / (len(events) * 2 * 10 + len(events_with_three_people) * 10) * 100)}")
    if model_score is not None:
        print(f"Total (with parameters) Score: {model_score}")
    else:
        print(f"Total (with parameters) Score: {total_score}")

    print(f"Execution time: {time.time() - start_time} seconds")

    if not NO_SAME_PAIRS:
        pair_to_events = {}
        for event, assigned_people in event_assignments.items():
            if len(assigned_people) == 2:
                pair = tuple(sorted(assigned_people))
                if pair not in pair_to_events:
                    pair_to_events[pair] = []
                pair_to_events[pair].append(event)
        for pair, events_list in pair_to_events.items():
            if len(events_list) > 1:
                print(f"{pair[0]} and {pair[1]} on {len(events_list)} events: {', '.join(events_list)}")

    # Calculate and print how many "should work together" pairs were actually assigned together

    output_target.close()

    workbook = load_workbook(filename=file_path)
    sheet_name = 'Output'

    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
    print(sheet)

    assignments = [[name] + subjects for name, subjects in assignments]
    assignments.sort()
    columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L",
               "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    for co in columns:
        for s in range(50):
            sheet[co + str(s + 1)] = ""
    sheet["A1"] = "Name"
    sheet["B1"] = "Event 1"
    sheet["C1"] = "Event 2"
    sheet["D1"] = "Event 3"

    for i in range(len(assignments)):
        for j in range(len(assignments[i])):
            square = columns[j] + str(i + 2)

            sheet[square] = assignments[i][j]

    sheet["F1"] = "Event Name"
    sheet["G1"] = "Person 1"
    sheet["H1"] = "Person 2"
    sheet["I1"] = "Person 3"

    event_assignments = [[subject] + names for subject, names in event_assignments.items()]

    for i in range(len(event_assignments)):
        for j in range(len(event_assignments[i])):
            square = columns[j + 5] + str(i + 2)

            sheet[square].value = event_assignments[i][j]

    try:
        workbook.save(file_path)

    except PermissionError:
        print_red("Unable to open " + str(file_path) + ", permission denied. Results are still in output.txt")


if __name__ == "__main__":
    solve()
    if print_logs:
        print(gc.get_stats())